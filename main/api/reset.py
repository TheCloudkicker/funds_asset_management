import json
import openpyxl
import xlrd
#DJANGO IMPORTS
from django.utils import timezone
from django.shortcuts import get_object_or_404, render, redirect, reverse

#DJANGO REST IMPORTS
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework import viewsets, status, generics, permissions
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny, IsAuthenticated


from entities.models import Fund, EntityOwnership, LegacyData, CapitalMovement
from fs.models import Fsli
from main.models import Currency, AuditSettings, Period
from supports.models import RepositoryType
from supports.models import Support, SupportType, ClientReport, ClientReportHeader


class ResetAPI(APIView):

    permission_classes = (IsAuthenticated,)

    def post(self, request, *args, **kwargs):

        method = request.data['method']

        if method == 'RESET':

            missing_ownerships = 0



            # PERIODS
            periods = Period.objects.all()
            for period in periods:
                period.delete()

            with open('fixtures/periods.json') as json_file:
                periods_list = json.load(json_file)
                for period_obj in periods_list:
                    period = Period(
                        pk=period_obj['pk'],
                        clientID=period_obj['fields']['clientID'],
                        start_date=period_obj['fields']['start_date'],
                        end_date=period_obj['fields']['end_date'],
                        active=period_obj['fields']['active'],
                    )
                    period.save()


            legacyCopy = list(LegacyData.objects.filter(fund_id=56).values())

            print('legacyCopy', legacyCopy)
            
            legacyDatas = LegacyData.objects.all()
            for legacyData in legacyDatas:
                legacyData.delete()




            movements_qs = CapitalMovement.objects.all()
            movements_copy = list(movements_qs.values())

            for movment in movements_qs:
                movment.delete()

            #FUNDS / ENTITY OWNERSHIPS
            funds = Fund.objects.all()
            for fund in funds:
                fund.delete()

            with open('fixtures/funds.json') as json_file:
                funds_list = json.load(json_file)
                for fund_obj in funds_list:
                    fund = Fund(
                        clientID=1,
                        pk=fund_obj['pk'],
                        prefRate=fund_obj['fields']['prefRate'],
                        legal_name=fund_obj['fields']['legal_name'],
                        trackingID=fund_obj['fields']['trackingID'],
                        isAudited=fund_obj['fields']['isAudited'],
                        commencement_date=fund_obj['fields']['commencement_date'],
                    )
                    fund.save()

            for movement in movements_copy:

                mov = CapitalMovement(
                    created_dt=movement['created_dt'],
                    uuid=movement['uuid'],
                    fund_id=movement['fund_id'],
                    movementType=movement['movementType'],
                    movementDate=movement['movementDate'],
                    totalAmount=movement['totalAmount'],
                    dateAdded=movement['dateAdded'],
                    supportID=movement['supportID']
                )
                mov.save()

            for lcopy in legacyCopy:

                l = LegacyData(
                    created_dt=lcopy['created_dt'],
                    fund_id=lcopy['fund_id'],
                    data=lcopy['data'],
                )
                l.save()

            
            entity_ownerships = EntityOwnership.objects.all()
            for entity_ownership in entity_ownerships:
                entity_ownership.delete()

            funds = Fund.objects.all()
            ownership_file = openpyxl.load_workbook('fixtures/legacy/ownerships.xlsx')
            ownership_sheet = ownership_file['mappings']


            for row in ownership_sheet.iter_rows():

                ownerID = row[0].value
                ownership_percentage = row[1].value
                owneeID = row[2].value

                try: owner = funds.get(trackingID=ownerID)
                except:
                    missing_ownerships += 1  
                    continue

                try: ownee = funds.get(trackingID=owneeID)
                except:
                    missing_ownerships += 1 
                    continue

                if owner and ownee:
                    entity_ownership = EntityOwnership(
                        owner_entity=owner,
                        owned_entity=ownee,
                        ownership_percentage=ownership_percentage,
                        ownership_type="SHARED"
                    )

                    entity_ownership.save()

            #REPORTS
            reports = ClientReport.objects.all()
            for report in reports:
                report.delete()

            headers = ClientReportHeader.objects.all()
            for header in headers:
                header.delete()

            repo_types = RepositoryType.objects.all()
            for repo_type in repo_types:
                repo_type.delete()

            fund_supports = Support.objects.all()
            for fund_support in fund_supports:
                fund_support.delete()

            support_types = SupportType.objects.all()
            for support_type in support_types:
                support_type.delete()


            with open('fixtures/repo_types.json') as json_file:
                repo_types_list = json.load(json_file)
                for repo_type_obj in repo_types_list:
                    repo_type = RepositoryType(
                        clientID=1,
                        pk=repo_type_obj['pk'],
                        name=repo_type_obj['fields']['name'],
                    )
                    repo_type.save()


            with open('fixtures/reports.json') as json_file:
                reports_list = json.load(json_file)
                for report_obj in reports_list:

                    repositoryType = RepositoryType.objects.get(pk=report_obj['fields']['repositoryType'])

                    report = ClientReport(
                        clientID=1,
                        name=report_obj['fields']['name'],
                        trackingInfo=report_obj['fields']['trackingInfo'],
                        repoType=repositoryType
                    )
                    report.save()

                    support_type = SupportType(
                        clientID=1,
                        name=report_obj['fields']['name']
                    )
                    support_type.save()

            with open('fixtures/client_headers.json') as json_file:
                headers_list = json.load(json_file)
                for header_obj in headers_list:
                    report = ClientReport.objects.get(name="Trial Balance")

                    header = ClientReportHeader(
                        clientID=1,
                        name=header_obj['fields']['name'],
                        report=report
                    )
                    header.save()


            # AUDIT SETTINGS

            audit_settings = AuditSettings.objects.all()
            for audit_setting in audit_settings:
                audit_setting.delete()

            with open('fixtures/audit_settings.json') as json_file:
                audit_settings_list = json.load(json_file)
                for audit_setting_obj in audit_settings_list:
                    audit_setting = AuditSettings(
                        pk=audit_setting_obj['pk'],
                        name=audit_setting_obj['fields']['name'],
                        benchmark=audit_setting_obj['fields']['benchmark'],
                        overall=audit_setting_obj['fields']['overall'],
                        performance=audit_setting_obj['fields']['performance'],
                        deminimis=audit_setting_obj['fields']['deminimis'],
                    )
                    audit_setting.save()

        headers_file = openpyxl.load_workbook('fixtures/legacy/investment_headers.xlsx')
        headers_sheet = headers_file['MQ Report']
        holdingsReport = ClientReport.objects.get(name="Holdings Report")

        for row in headers_sheet.iter_rows():
            for cell in row:
                if cell.value:
                    value = cell.value.replace(" ", "")
                    header = ClientReportHeader(
                        clientID=1,
                        name=value,
                        report=holdingsReport
                    )
                    header.save()


        cdr_workbook = xlrd.open_workbook('fixtures/legacy/cdr.xls')
        capitalReport = ClientReport.objects.get(name="Capital Report")
        cdr_sheet = cdr_workbook.sheet_by_index(0)
        row = cdr_sheet.row(1)

        all_headers = []
        unique_headers = []

        for _, cell_obj in enumerate(row):

            name = cell_obj.value

            if name in unique_headers:
                waibe_name = f'{name}_2'
            else:
                waibe_name = name
                unique_headers.append(name)

            obj = {
                'name':name,
                'waibe_name':waibe_name
            }

            all_headers.append(obj)

        for header_obj in all_headers:
            header = ClientReportHeader(
                    clientID=1,
                    name=header_obj['name'],
                    waibe_name=header_obj['waibe_name'],
                    report=capitalReport
                )
            header.save()




            # fund_criterions = FundCriterionHeader.objects.all()
            # for fund_criterion in fund_criterions:
            #     fund_criterion.delete()

            # fslis = Fsli.objects.all()
            # for fsli in fslis:
            #     fsli.delete()


                    


  


            # currencies = Currency.objects.all()
            # for currency in currencies:
            #     currency.delete()

            # with open('fixtures/currencies.json') as json_file:
            #     ccys_list = json.load(json_file)
            #     for ccy_obj in ccys_list:
            #         ccy = Currency(
            #             short_name=ccy_obj['fields']['short_name'],
            #             symbol=ccy_obj['fields']['symbol'],
            #         )
            #         ccy.save()


            response = {
                'missing_ownerships':missing_ownerships
            }

            return Response(response,  status=status.HTTP_200_OK)


